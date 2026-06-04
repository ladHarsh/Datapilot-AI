"""
Excel Export Service
====================
Exports SQL query results to formatted Excel (.xlsx) workbooks.
Uses openpyxl for formatting with styled headers and auto-sized columns.
Validates export size limits to prevent memory overload.
"""

import io
import logging
from typing import List, Dict, Any, Optional

from fastapi.responses import StreamingResponse

logger = logging.getLogger(__name__)

# ─── Export limits ────────────────────────────────────────────────────────────
MAX_EXPORT_ROWS = 50_000
MAX_EXPORT_SIZE_BYTES = 50 * 1024 * 1024   # 50 MB


class ExcelExportService:
    """
    Converts query result rows into a downloadable Excel (.xlsx) workbook.
    Supports styled headers, auto-column widths, and multiple sheets.
    """

    def export(
        self,
        rows: List[Dict[str, Any]],
        user_role: str,
        filename: str = "query_results.xlsx",
        sheet_name: str = "Results",
        max_rows: int = MAX_EXPORT_ROWS,
        include_row_numbers: bool = False,
    ) -> StreamingResponse:
        """
        Generate an Excel StreamingResponse from query result rows.

        Args:
            rows:               Query result rows as list of dicts.
            user_role:          Role of the user requesting the export.
            filename:           Desired download filename.
            sheet_name:         Name of the Excel worksheet.
            max_rows:           Max rows to export.
            include_row_numbers: Add a row number column.

        Returns:
            FastAPI StreamingResponse with Excel content.
        """
        logger.info("ExcelExportService: export requested by role='%s'.", user_role)

        # 1. Permission check
        if user_role.lower() not in ["admin", "analyst"]:
            logger.warning("ExcelExportService: unauthorized export attempt by role='%s'.", user_role)
            raise PermissionError(f"Role '{user_role}' is not authorized to export data. Only admins and analysts can export.")

        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError(
                "openpyxl is required for Excel export. "
                "Install it with: pip install openpyxl"
            )

        logger.info("ExcelExportService: exporting %d rows.", len(rows))

        # Row limit
        effective_limit = min(max_rows, MAX_EXPORT_ROWS)
        truncated = False
        if len(rows) > effective_limit:
            logger.warning(
                "ExcelExportService: truncating from %d to %d rows.",
                len(rows), effective_limit,
            )
            rows = rows[:effective_limit]
            truncated = True

        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = sheet_name[:31]   # Excel sheet name limit

        if not rows:
            ws.append(["No data available."])
        else:
            headers = list(rows[0].keys())
            if include_row_numbers:
                headers = ["#"] + headers

            # ── Header row styling ────────────────────────────────────────────
            header_fill = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
            header_font = Font(color="FFFFFF", bold=True, name="Calibri", size=11)
            header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            thin_border = Border(
                left=Side(style="thin"),
                right=Side(style="thin"),
                top=Side(style="thin"),
                bottom=Side(style="thin"),
            )

            ws.append(headers)
            for col_idx, cell in enumerate(ws[1], start=1):
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                cell.border = thin_border

            ws.row_dimensions[1].height = 20

            # ── Data rows ────────────────────────────────────────────────────
            alt_fill = PatternFill(start_color="EBF3FB", end_color="EBF3FB", fill_type="solid")
            data_font = Font(name="Calibri", size=10)
            data_alignment = Alignment(vertical="center", wrap_text=False)

            for row_num, row in enumerate(rows, start=1):
                row_values = list(row.values())
                if include_row_numbers:
                    row_values = [row_num] + row_values
                ws.append(row_values)

                # Alternate row shading
                if row_num % 2 == 0:
                    for cell in ws[row_num + 1]:  # +1 for header offset
                        cell.fill = alt_fill

                for cell in ws[row_num + 1]:
                    cell.font = data_font
                    cell.alignment = data_alignment
                    cell.border = thin_border

            # ── Auto-size columns ─────────────────────────────────────────────
            for col_idx, col_header in enumerate(headers, start=1):
                col_letter = get_column_letter(col_idx)
                max_len = len(str(col_header))
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        try:
                            max_len = max(max_len, len(str(cell.value or "")))
                        except Exception:
                            pass
                ws.column_dimensions[col_letter].width = min(max_len + 4, 50)

            # Freeze header row
            ws.freeze_panes = "A2"

        # ── Add metadata sheet if truncated ───────────────────────────────────
        if truncated:
            meta_ws = wb.create_sheet("Export Info")
            meta_ws.append(["Export Metadata"])
            meta_ws.append(["Rows Exported", len(rows)])
            meta_ws.append(["Truncated", "Yes — export limit reached"])
            meta_ws.append(["Max Rows Limit", effective_limit])

        # Serialize to bytes
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.getvalue()

        # Size guard
        if len(content) > MAX_EXPORT_SIZE_BYTES:
            raise ValueError(
                f"Excel export size ({len(content) / 1024 / 1024:.1f} MB) "
                f"exceeds the maximum allowed ({MAX_EXPORT_SIZE_BYTES / 1024 / 1024:.0f} MB)."
            )

        logger.info(
            "ExcelExportService: generated %.2f KB%s.",
            len(content) / 1024,
            " (truncated)" if truncated else "",
        )

        safe_filename = self._sanitize_filename(filename)

        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_filename}"',
                "X-Row-Count": str(len(rows)),
                "X-Truncated": str(truncated).lower(),
                "X-Export-Size-Bytes": str(len(content)),
            },
        )

    def export_multi_sheet(
        self,
        sheets: Dict[str, List[Dict[str, Any]]],
        filename: str = "multi_sheet_export.xlsx",
        max_rows_per_sheet: int = MAX_EXPORT_ROWS,
    ) -> StreamingResponse:
        """
        Export multiple query result sets into separate Excel sheets.

        Args:
            sheets: Dict mapping sheet name → list of row dicts.
            filename: Output filename.
            max_rows_per_sheet: Row limit per sheet.

        Returns:
            StreamingResponse with multi-sheet Excel file.
        """
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils import get_column_letter
        except ImportError:
            raise ImportError("openpyxl is required for Excel export.")

        wb = openpyxl.Workbook()
        wb.remove(wb.active)   # Remove default empty sheet

        for sheet_name, rows in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])
            rows = rows[:max_rows_per_sheet]

            if rows:
                headers = list(rows[0].keys())
                ws.append(headers)
                for row in rows:
                    ws.append(list(row.values()))

        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        content = buffer.getvalue()

        safe_filename = self._sanitize_filename(filename)

        return StreamingResponse(
            iter([content]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{safe_filename}"',
                "X-Sheet-Count": str(len(sheets)),
            },
        )

    @staticmethod
    def _sanitize_filename(filename: str) -> str:
        import re
        safe = re.sub(r"[^\w\-. ]", "_", filename)
        return safe[:200]
